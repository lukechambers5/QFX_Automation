import tkinter as tk
from tkinter import filedialog, messagebox
import tkinter.font as tkFont
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import numbers
import numpy as np

def abbreviate(df):
    replacements = {
        "Concord Medical Group": "CMG",
        "Great Lakes Emergency": "GLEP",
        "South Central Physicians": "SCP",
        "Mid West Hospital Phys": "MWHP",
        "CMG of KY": "CMGofKY",
        "Four Corners Emergency": "FCEP",
        "Western Mountain Hospital": "WMHP",
        "Concord Company of Tennessee": "CCofTN",
        "Concord North Texas": "CNT",
        "Delaware River Medicine": "DRM"
    }
    if "ACCOUNT_NAME" in df.columns:
        for old_value, new_value in replacements.items():
            df["ACCOUNT_NAME"] = df["ACCOUNT_NAME"].apply(
                lambda x: x.replace(old_value, new_value) if isinstance(x, str) and old_value.lower() in x.lower() else x
            )
    return df

def delete_columns(df):
    columns_to_delete = ["DFI_ID", "ACCOUNT_NUMBER", "FITID", "CHECK_NO"]
    return df.drop(columns=[col for col in columns_to_delete if col in df.columns])

def dates(df):
    if 'DTPOSTED' in df.columns:
        date = pd.to_datetime(df['DTPOSTED'].str[:8], format='%Y%m%d', errors='coerce') \
                    .dt.strftime('%m/%d/%Y')
        df.insert(0, "DATE", date)
    else:
        df.insert(0, "DATE", pd.Series([""] * len(df), dtype=str))
    return df

def export_to_excel_with_text_format(df, output_path):
    wb = Workbook()
    ws = wb.active
    # Write DataFrame to worksheet
    for row in dataframe_to_rows(df, index=False, header=True):
        if row[0] is None or (isinstance(row[0], float) and np.isnan(row[0])):
            continue
        ws.append(row)
    
    # Set DATE column (A) format as text to preserve 06/06/2025
    for cell in ws['A']:
        cell.number_format = numbers.FORMAT_TEXT  # equivalent to "@"

    wb.save(output_path)

def create_id(df):
    def build_id(row):
        if pd.isna(row["ACCOUNT_NAME"]) or pd.isna(row["DATE"]) or pd.isna(row["TRNAMT"]):
            return None
        try:
            date = pd.to_datetime(row["DATE"])
            excel_serial = (date - pd.Timestamp("1899-12-30")).days
            return f"{str(row['ACCOUNT_NAME']).strip()}{excel_serial}{str(row['TRNAMT']).strip()}"
        except:
            return None
    
    df.insert(0, "ID", df.apply(build_id, axis=1))
    return df

def handle_duplicates(df, output_path):
    duplicates = df[df.duplicated(subset=["ID"], keep=False)]
    non_duplicates = df.drop(duplicates.index)
    
    duplicates = create_abbreviation(duplicates)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Cleaned Data"
    for row in dataframe_to_rows(non_duplicates, index=False, header=True):
        if all((cell is None or (isinstance(cell, float) and np.isnan(cell))) for cell in row):
            continue
        ws1.append(row)
    for cell in ws1['A']:
        cell.number_format = numbers.FORMAT_TEXT

    ws2 = wb.create_sheet(title="Duplicates")
    for row in dataframe_to_rows(duplicates, index=False, header=True):
        if row[0] is None or (isinstance(row[0], float) and np.isnan(row[0])):
            continue
        ws2.append(row)
    for cell in ws2['A']:
        cell.number_format = numbers.FORMAT_TEXT

    split_double_duplicates(wb)

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        delete_column_from_sheet(ws, "DTPOSTED")

    wb.save(output_path)

def create_abbreviation(df):
    if "ID" in df.columns and "NAME" in df.columns:
        df["ID"] = df.apply(
            lambda row: str(row["ID"]) + str(row["NAME"])[:5] if pd.notna(row["ID"]) and pd.notna(row["NAME"]) else row["ID"],
            axis=1
        )
    return df

def split_double_duplicates(wb):
    ws_duplicates = wb["Duplicates"]

    data = list(ws_duplicates.iter_rows(values_only=True))
    if not data:
        return

    headers = data[0]
    rows = data[1:]

    df_dup = pd.DataFrame(rows, columns=headers)

    double_dups = df_dup[df_dup.duplicated(subset=["ID"], keep=False)]
    cleaned_duplicates = df_dup.drop(double_dups.index)

    ws_duplicates.delete_rows(2, ws_duplicates.max_row)

    for row in dataframe_to_rows(cleaned_duplicates, index=False, header=False):
        if all(cell is None or (isinstance(cell, float) and np.isnan(cell)) for cell in row):
            continue
        ws_duplicates.append(row)

    ws3 = wb.create_sheet(title="Double Duplicates")
    for row in dataframe_to_rows(double_dups, index=False, header=True):
        if all(cell is None or (isinstance(cell, float) and np.isnan(cell)) for cell in row):
            continue
        ws3.append(row)

    for cell in ws_duplicates['A']:
        cell.number_format = numbers.FORMAT_TEXT
    for cell in ws3['A']:
        cell.number_format = numbers.FORMAT_TEXT

def delete_column_from_sheet(ws, col_name):
    headers = [cell.value for cell in ws[1]]
    if col_name in headers:
        col_idx = headers.index(col_name) + 1 
        ws.delete_cols(col_idx)

def process_file(file_path):
    try:
        status_label.config(text="Processing...", fg="orange")
        root.update_idletasks()

        df = pd.read_csv(file_path, dtype=str)
        df = abbreviate(df)
        df = delete_columns(df)
        df = dates(df)
        df = create_id(df)

        # Changes from CSV to XLSX
        input_directory = os.path.dirname(file_path)
        base_filename = "PROCESSED_" + os.path.basename(file_path).replace(".csv", ".xlsx")
        output_full_path = os.path.join(input_directory, base_filename)

        handle_duplicates(df, output_full_path)
        select_button.grid_remove() # Hide the select button
        open_file_button.config(command=lambda: open_processed_file(output_full_path))
        open_file_button.grid(row=0, column=0, padx=5, pady=5, sticky="") # Show the open file button

        status_label.config(text="Processing Complete!", fg="green")
    except Exception as e:
        status_label.config(text="Failed!", fg="red")
        messagebox.showerror("Error", f"Failed to process file:\n{e}")

def reset_ui():
    status_label.config(text="", fg="#CCCCCC")
    open_file_button.grid_remove() # Hide the open file button
    select_button.grid(row=0, column=0, padx=5, pady=5, sticky="") 

def select_file():
    file_path = filedialog.askopenfilename(title="Select a file", filetypes=[("CSV Files", "*.csv")])
    if file_path:
        process_file(file_path)

def start_move(event):
    root.x = event.x
    root.y = event.y

def stop_move(event):
    root.x = None
    root.y = None

def do_move(event):
    deltax = event.x - root.x
    deltay = event.y - root.y
    x = root.winfo_x() + deltax
    y = root.winfo_y() + deltay
    root.geometry(f"+{x}+{y}")

def open_processed_file(path):
    try:
        os.startfile(path)  # Windows only
    except Exception as e:
        messagebox.showerror("Error", f"Could not open file:\n{e}")

def on_enter(e):
    select_button['bg'] = "#45a049"

def on_leave(e):
    select_button['bg'] = "#4CAF50"

def on_enter_open(e):
    open_file_button['bg'] = "#45a049"

def on_leave_open(e):
    open_file_button['bg'] = "#4CAF50"

root = tk.Tk()
root.title("File Processor")
root.resizable(False, False)

accent_color = "#4CAF50"

# Bindings for window movement
root.bind("<ButtonPress-1>", start_move)
root.bind("<ButtonRelease-1>", stop_move)
root.bind("<B1-Motion>", do_move)

status_label = tk.Label(root, text="", font=("Segoe UI", 9), fg="#CCCCCC")
status_label.grid(row=2, column=0, pady=(0,0)) # Give some padding from the button

window_width = 500
window_height = 350
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x = int((screen_width / 2) - (window_width / 2))
y = int((screen_height / 2) - (window_height / 2))
root.geometry(f"{window_width}x{window_height}+{x}+{y}")

custom_font = tkFont.Font(family="Segoe UI", size=11, weight="bold")

# Create a frame to hold the button for better centering control
button_frame = tk.Frame(root)
button_frame.grid(row=0, column=0, sticky="nsew")

# Configure grid weights for the frame to make it expandable and center content
button_frame.grid_rowconfigure(0, weight=1)
button_frame.grid_columnconfigure(0, weight=1)

# Create the select button
select_button = tk.Button(button_frame, text="Select CSV File", command=select_file, font=custom_font, bg=accent_color, fg="white", padx=15, pady=8)
select_button.grid(row=0, column=0, padx=5, pady=5, sticky="")

select_button.bind("<Enter>", on_enter)
select_button.bind("<Leave>", on_leave)
select_button.config(relief="flat")

# Create the open file button but hide it initially
open_file_button = tk.Button(button_frame, text="Open Processed File", font=custom_font, bg=accent_color, fg="white", padx=15, pady=8, relief="flat")
open_file_button.grid_remove()

open_file_button.bind("<Enter>", on_enter_open)
open_file_button.bind("<Leave>", on_leave_open)

# Configure root's grid to make the button_frame and status_label fill the window
root.grid_rowconfigure(0, weight=1) # Row for the button frame
root.grid_rowconfigure(1, weight=0) # Row for the status label (don't expand)
root.grid_columnconfigure(0, weight=1) # Column for all content

root.mainloop()
